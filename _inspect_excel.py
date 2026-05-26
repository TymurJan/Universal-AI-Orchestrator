"""Інспектор структури Excel-шаблону бюджету IREX."""
import openpyxl

EXCEL_PATH = r"D:\ГО Талан UA\Гранти\Конкурс IREX\2. Project Budget_Бюджет проєкту_шаблон_upd_VK_11.25 - Copy.xlsm"

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, keep_vba=True)
print(f"\n=== Sheets: {wb.sheetnames} ===\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"АРКУШ: {sheet_name}  (rows: {ws.max_row}, cols: {ws.max_column})")
    print(f"{'='*60}")
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Print non-empty rows only
        row_data = [str(c) if c is not None else "" for c in row]
        if any(c.strip() for c in row_data):
            print(f"  R{row_idx:03d}: {row_data}")
        if row_idx > 200:
            print("  ... (truncated after 200 rows)")
            break

wb.close()
print("\nDone.")
